#!/usr/bin/env python3
"""
Author : Jon Freivald <jfreivald@brmedical.com>
        : Copyright © Blue Ridge Medical Center, 2024. All Rights Reserved.
        : License: GNU GPL Version 3
Date   : 2024-06-05
Purpose: NPI extraction tool
        : Uses Excel spreadsheets as data sources.
        : Extracts NPIs from master Medicaid spreadsheet (>150,000 records) based on local list of provider NPIs.
        : Version change log at EoF.
"""

# Distribution license for PiSimpleGUI v5
PySimpleGUI_License = 'eYysJ6MYaPWpN6lzbMnONmlIVUHCllwoZdSTIF63ITklRppfcS3VRMyja6WiJY1VdRG1lvvvbSiUI7sVIKkPxEpTYw2WVIuocj2yVdJwR5C2IJ6uMoTIcQxCOWDHgJxdNXTSQy5xMKixw7icTeGrlHjUZvW25Ez4ZvUYRelwcwGoxqvBe5Wz1alYbFnzR9WsZ2XdJLz8aoWx9uu0IUjtoLivNNSn40wHIti5wniDTRmhF7tMZYU1ZqpKcMnbNr0QIZjUoBizSzmE9fuOIyiSwIiOTTmAFmtoZeUOxsh8cL3SQCiMO6iGJkGhcLmIVipIdYmyFrszZDC3IHsNI2kzN9vCboX2BAhAb2nGkzidOnioJiClbdHgVzlIIDFQJSpzZOGWd2lRItEy1Pl2ZQGBlmjVY4Www8g1QK2tV2updCGcVmySIVibw1ikQ73MVuzXdGG89ttyZdXlJZJKR1CNIa6wIbjKkS39NrTyEGiqLcCnJOEOYSXRRWlWSOXqNOzqd7WvVnknIHjHoIi9MXjdAmyrNGCl0HwCN3im0OxgO4SBICstIfk2RHhgdsGGVOFAeyHmBopscGm5V6zXIlj9owi1MijPAby7NTSB0qwBMayP0rxtOxSMI6sLIHk2VPtNYkWYldsWQTW9RukFcomHVWzlcryLIY6rImmJpLmSc1mDVUpUdambFSsaZME2BriWchmg1yliZCGhlfjbYFWDwIuiYy2E9atuIFiSwwiLSaVeBPBzZUGURlyeZJXENtzgINjBo3igMejAAz1TLIjzIGyVMbCZ40yWMDzHMEu2MfTcEt2yIXnP0G=5160c5e7f3a0b230efb3661d7a3268ddfd43a0cd8fd74c4308c51bf55ac37cc8544863bed7e251e0e551821240fbb7a92d5f811e068091dc3a337b4743ee4e09b729146f0924a668c7664c0acfec7785021b6ea12e45bb61134318db27191eeeae4b00e2c9ad1894f1e3f69499594997119fa53d53841b3047e9f060510cb405b331b18fff23685bbc3715bd31462f9e8ff6dd1408786a939241b376fb69c3636203419f14648708c96a651cb862eba7244862a76dca8782ddb333f0a31b66eaeb37349e0bfdda847430f4210b7fdd6f9d667ec12effa82ba76035f99f0cad2ecbc7285a10fdd1c0e9bdf831aef6c14532163d2d6c18ab53a0f2d75e828d433b3040adc67346f64048b1851e0d2f396316e2476259f00083826a12a2ea1ff2cfeec1b8e3b4e2d3ee52c24a824a608e5d04f1b3e8539f0a27b19cacd0d0619519b2d493f6eda6a1f404be388d20ae237518d0119a214b6b9e3e4e8e25ff655c9cee8d1e658127af5a38f36fcbe13be1f48b4bf2faca18ea7cfa621d22de232533ba86b9ca392c64770644cd5bb6fb95e45e2c1894b187fc890a77a41892f18cbc43d7861dc79a1426efce3a1ca9210ad3b848ed5771924e479290becf7a315776a0f31abcb46baacfe318209f0a31c68d10d6476609469e096ea89f1d8e787ede1f34cd1712cb2b3fabb3a99864c368b659762f77f50752836749f86eabf85dd08'

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import argparse
import PySimpleGUI as sg
import os
import pickle
import time
import subprocess
from datetime import datetime

BRMC = {'BACKGROUND': '#73afb6',
                 'TEXT': '#00446a',
                 'INPUT': '#ffcf01',
                 'TEXT_INPUT': '#00446a',
                 'SCROLL': '#ce7067',
                 'BUTTON': ('#ffcf01', '#00446a'),
                 'PROGRESS': ('#ffcf01', '#00446a'),
                 'BORDER': 1, 'SLIDER_DEPTH': 0, 'PROGRESS_DEPTH': 0,
                 }
sg.theme_add_new('BRMC', BRMC)

progver = 'v 1.0'
mainTheme = 'BRMC'
errorTheme = 'HotDogStand'
config_file = (f'{os.path.expanduser("~")}/npi_config.dat')
output_file = (f'{os.path.expanduser("~")}/Downloads/MedicaidProviderExtract.xlsx')

# --------------------------------------------------
def get_args():
	"""Process any command line arguments"""

	parser = argparse.ArgumentParser(
		description='View provider Medicaid validation status.',
		formatter_class=argparse.ArgumentDefaultsHelpFormatter)

	parser.add_argument('-f',
		'--file',
		help='Local provider list containing NPIs to look up.',
		metavar='filename',
		type=str,
		default='Providers.xlsx')    # Put your default file here! NPI in column D, data starts in row 8. 
                                    # (or update the code to match your format)

	args = parser.parse_args()

	return args

# --------------------------------------------------
def get_user_settings():

    user_config = {}

    try:
        with open(config_file, 'rb') as fp:
            user_config = pickle.load(fp)
        fp.close()
    except:
        user_config['Theme'] = mainTheme

    return user_config

# --------------------------------------------------
def write_user_settings(user_config):

    try:
        with open(config_file, 'wb') as fp:
            pickle.dump(user_config, fp)
        fp.close()
    except:
        sg.theme(errorTheme)
        layout = [ [sg.Text(f'File or data error: {config_file}. Updates NOT saved!')],
					[sg.Button('Quit')] ]
        window = sg.Window('FILE ERROR!', layout, finalize=True)

        while True:
            event, values = window.read()
            if event in (sg.WIN_CLOSED, 'Quit'): # if user closes window or clicks quit
                window.close()
                return

# --------------------------------------------------
def get_part_of_day(h):
    return (
        "morning"
        if h <= 11
        else "afternoon"
        if 12 <= h <= 17
        else "evening"
        if 18 <= h <= 22
        else "night"
    )
# --------------------------------------------------
def create_extract(output_file, args, window):

    window['-STATUS_MSG-'].update('Loading local NPIs...')
    window.refresh()

    npiList = [] # Provider NPI List
    try:
        npiFile = openpyxl.load_workbook(args.file)
    except:
        sg.theme(errorTheme)
        layout = [ [sg.Text(f'Unable to open {args.file} to load NPI data. Unable to continue')],
					[sg.Button('Quit')] ]
        window = sg.Window('FILE ERROR!', layout, finalize=True)

        while True:
            event, values = window.read()
            if event in (sg.WIN_CLOSED, 'Quit'): # if user closes window or clicks quit
                window.close()
                return False

    currentSheet = npiFile['Providers']
    currentProvider = ''
    for row in range(8, currentSheet.max_row + 1): # NPI data starts in row 8
        cellA = (f'A{row}')
        cellD = (f'D{row}')
        if currentSheet[cellA].value != None: # We were pulling blank rows for some reason -- filter them out!
            if currentSheet[cellA].value != currentProvider:
                currentProvider = currentSheet[cellA].value
            currentNPI = currentSheet[cellD].value
            npiList.append([currentNPI])

    npiFile.close()

    window['-STATUS_MSG-'].update(f'{len(npiList)} local NPIs loaded.')
    window.refresh()

    try:
        file = sg.popup_get_file('Select the Medicaid source file.', title='Source data')
        window['-STATUS_MSG-'].update(f'Loading Medicaid source file. BE PATIENT!', text_color='red')
        window.refresh()
        start = time.perf_counter()
        src = openpyxl.load_workbook(file)
    except:
        return False
    
    currentSheet = src.active

    num_rows = 0

    window['-STATUS_MSG-'].update(f'Processing {num_rows} records...', text_color='#00446a')
    window.refresh()

    wb = Workbook() # create our output file
    ws = wb.active

    ws['A1'] = 'NPI'
    ws['B1'] = 'Last Name'
    ws['C1'] = 'First Name'
    ws['D1'] = 'Middle Name'
    ws['E1'] = 'Enrollment Type'
    ws['F1'] = 'Revalidation Date'
    ws['G1'] = 'Provider Type Description'
    ws['H1'] = 'Zip Code'

    for row in range(2, currentSheet.max_row + 1): # our file has headers
        c_npi = (f'B{row}')
        c_lname = (f'E{row}')
        c_fname = (f'C{row}')
        c_mi = (f'D{row}')
        c_en_type = (f'K{row}')
        c_rv_date = (f'M{row}')
        c_type_desc = (f'H{row}')
        c_zip_code = (f'L{row}')

        if currentSheet[c_npi].value != None: # eliminate blank lines
            num_rows += 1
            for i in npiList:
                if currentSheet[c_npi].value in i:
                    npi = currentSheet[c_npi].value
                    lname = currentSheet[c_lname].value
                    fname = currentSheet[c_fname].value
                    mi = currentSheet[c_mi].value
                    en_type = currentSheet[c_en_type].value
                    rv_date = currentSheet[c_rv_date].value
                    type_desc = currentSheet[c_type_desc].value
                    zip_code = currentSheet[c_zip_code].value
                    ws.append([npi, lname, fname, mi, en_type, rv_date, type_desc, zip_code])
            if num_rows % 50 == 0:
                window['-STATUS_MSG-'].update(f'Processing {num_rows} records...')
                window.refresh()

    for cell in ws['F']:
        cell.number_format = 'YYYY-MM-DD;@'

    for cell in ws['H']:
        cell.number_format = '00000-0000'

    tab = Table(displayName='Providers', ref=f'A1:H{ws.max_row}')
    style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    end = time.perf_counter()
    window['-STATUS_MSG-'].update(f'Done! {ws.max_row -1} records extracted from {num_rows} in {(round(end-start))} seconds.')
    window.refresh()

    try:
        wb.save(output_file) # write the output file
    except:
        sg.theme(errorTheme)
        layout = [ [sg.Text(f'File or data error: {output_file}. Updates NOT saved!')],
					[sg.Button('Quit')] ]
        window = sg.Window('FILE ERROR!', layout, finalize=True)

        while True:
            event, values = window.read()
            if event in (sg.WIN_CLOSED, 'Quit'): # if user closes window or clicks quit
                window.close()
                return False
    return True

# --------------------------------------------------
def extract_NPI_data():

    args = get_args()
    user_config = get_user_settings()
    if 'winLoc' in user_config:
        winLoc = user_config['winLoc']
    else:
        winLoc = (50, 50)
    if 'winSize' in user_config:
        winSize = user_config['winSize']
    else:
        winSize = (450, 260)

    part = get_part_of_day(datetime.now().hour)

    name = subprocess.check_output(
        'net user "%USERNAME%" /domain | find /I "Full Name"', shell=True, text=True
    )
    full_name = name.replace("Full Name", "").strip()
    first_name = full_name.split()[0]
    
    
    sg.theme(user_config['Theme'])
    layout = [  [sg.Image('logo.png', size=(400, 96))],
                [sg.Text(f'Good {part}, {first_name}')],
                [sg.Text('Extract local NPI data from Medicaid source file.')],
                [sg.Text('', key='-STATUS_MSG-')],
                [sg.Button('Open'), sg.Text('<-- Medicaid source file'), sg.Push(), sg.Button('Quit')],
                 [sg.Push(), sg.Text('Copyright © Blue Ridge Medical Center, 2024')] ]
    window = sg.Window(f'Provider NPI Query Tool {progver}', layout, location=winLoc, size=winSize, element_justification='center', grab_anywhere=True, resizable=True, finalize=True)
    window.BringToFront()

    while True:
        
        event, values = window.read(timeout=5000)
        winLoc = window.CurrentLocation()
        winSize = window.Size

        if event in (sg.WIN_CLOSED, 'Quit'): # if user closes window or clicks cancel
            if event == 'Quit':
                user_config['winLoc'] = winLoc
                user_config['winSize'] = winSize
                write_user_settings(user_config)
            break

        if event == 'Open':
            if create_extract(output_file, args, window):
                user_config['winLoc'] = winLoc
                user_config['winSize'] = winSize
                write_user_settings(user_config)
                sg.popup(f'Extraction complete. Your data is in {output_file}', title='Success!')
                subprocess.Popen(f"cmd /c start excel {output_file}")
            break
                
    window.close()

# --------------------------------------------------
if __name__ == '__main__':
	extract_NPI_data()
        
"""Change log:

    v 0.1   : 240605    : Initial version
    v 0.2   : 240606    : Added status update messages
    v 0.3   : 240606    : Just can't stop tweaking... Minor layout & message changes. Saves user settings when run now instead
                        : of just on "Quit"
    v 0.4   : 240611    : Updated to give a running record count as the source file is processed (for "not hung" feedback!)
    v 0.5   : 240612    : Additional UI update with more user feedback & a process timer.
    v 0.6   : 240612    : Updated file handling to include error checking.
    v 0.7   : 240612    : Updated output file to have data in an Excel table, and to apply desired formatting to date/zip cells.
    v 0.8   : 240619    : Was finally able to generate distrubution key for this app.
            : 240625    : Cleaned up a few comments and text -- no code changes.
    v 0.9   : 240913    : Added user greeting.
    v 1.0   : 241001    : Stable enough to be a 1.0 release. Updated to launch target file on exit.
"""